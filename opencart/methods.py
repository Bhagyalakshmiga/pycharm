import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
driver = webdriver.Chrome(ChromeDriverManager().install())
def registeropencart(fname, lname, email, password):
    driver.get("https://demo.opencart.com/")
    ActualTitle = driver.title
    ExpectedTitle = "Your Store"
    if ActualTitle == ExpectedTitle:
        print("Title Verified")
    else:
        print("Title Not Matched")
    driver.find_element(By.LINK_TEXT, "My Account").click()
    time.sleep(1)
    driver.find_element(By.LINK_TEXT, "Register").click()
    if driver.find_element(By.XPATH, "//*[@id='content']/h1").is_displayed():
        print("Heading verified")
    else:
        print("Heading not verified")
    time.sleep(2)
    driver.find_element(By.ID, "input-firstname").send_keys(fname)
    driver.find_element(By.ID, "input-lastname").send_keys(lname)
    driver.find_element(By.ID, "input-email").send_keys(email)
    driver.find_element(By.ID, "input-password").send_keys(password)
    driver.find_element(By.XPATH, "//input[@type='checkbox']").click()
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(2)

excelpath= "C:\\dataregist.xlsx"
workbook=openpyxl.load_workbook(excelpath)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
for i in range(2, rows+1):
    firstname=sheet.cell(row=i,column=1).value
    lastname=sheet.cell(row=i,column=2).value
    email=sheet.cell(row=i,column=3).value
    pword=sheet.cell(row=i,column=4).value
    registeropencart(firstname,lastname,email,pword)
    print("Registration successful")
