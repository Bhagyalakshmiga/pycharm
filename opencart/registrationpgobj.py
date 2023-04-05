import openpyxl
import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from opencart.pageobjects_reg import registration


@pytest.fixture()
def setup():
    print("This runs before method")

def test_def1(setup):
    print("Hello")

def test_registration(setup):
    excelpath = "C:\\dataregist.xlsx"
    workbook = openpyxl.load_workbook(excelpath)
    sheet = workbook.active
    rows = sheet.max_row
    cols = sheet.max_column
    print(rows)
    for i in range(2, rows + 1):
        firstname = sheet.cell(row=i, column=1).value
        lastname = sheet.cell(row=i, column=2).value
        email = sheet.cell(row=i, column=3).value
        pword = sheet.cell(row=i, column=4).value
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://demo.opencart.com/")
        rp= registration(driver)
        rp.clickonmyaccountlink()
        rp.clickonregistrationlink()
        rp.clickonenterfirstname(firstname)
        rp.clickonenterlastname(lastname)
        rp.clickonenterusername(email)
        rp.clickonenterpassword(pword)
        rp.clickoncheckbox()
        rp.clickonsubmitbutton()