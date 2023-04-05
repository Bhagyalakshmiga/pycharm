import openpyxl
import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from opencart.pageobjects import loginpage1
driver=webdriver.Chrome(ChromeDriverManager().install())

@pytest.fixture()

def setup():
    print("After every method")

def test_loginopencart1(setup):
            excelpath= "C:\\datap.xlsx"
            workbook=openpyxl.load_workbook(excelpath)
            sheet=workbook.active
            rows=sheet.max_row
            columns=sheet.max_column
            print(rows)
            for i in range(2,rows+1):
                uname=sheet.cell(row=i,column=1).value
                pword=sheet.cell(row=i,column=2).value
                driver=webdriver.Chrome(ChromeDriverManager().install())
                driver.get("https://demo.opencart.com/")
                lp=loginpage1(driver)
                lp.clickonmyaccount()
                lp.clickonloginlink()
                lp.clickonenteremail(uname)
                lp.clickonenterpassword(pword)
                lp.clickonsubmit()


