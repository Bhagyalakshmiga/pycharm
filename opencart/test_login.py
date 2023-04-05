import time
import openpyxl
import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
driver=webdriver.Chrome(ChromeDriverManager().install())
class test_Tc01:
        def loginopencart(username,password):
            driver.get("https://demo.opencart.com/")
            driver.find_element(By.LINK_TEXT,"My Account").click()
            time.sleep(5)
            driver.find_element(By.LINK_TEXT,"Login").click()
            driver.find_element(By.ID,"input-email").send_keys(username)
            driver.find_element(By.ID,"input-password").send_keys(password)
            driver.find_element(By.XPATH ,"//button[@type='submit']").click()
        excelpath= "C:\\datap.xlsx"
        workbook=openpyxl.load_workbook(excelpath)
        sheet=workbook.active
        rows=sheet.max_row
        cols=sheet.max_column
        print(rows)
        for i in range(2,rows+1):
            uname=sheet.cell(row=i,column=1).value
            pword=sheet.cell(row=i,column=2).value
            loginopencart(uname,pword)
print("login successful")
